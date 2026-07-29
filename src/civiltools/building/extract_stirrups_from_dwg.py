"""Extract stirrup (shear reinforcement) information from AutoCAD drawings via COM.

Supported text formats
----------------------
**Stirrup spacing** (single text object):
    Size / Spacing :  ``T12@15``  or  ``∅12@15``

**Stirrup ADD** (additional stirrup, single text):
    ``1T12(ADD)``  or  ``1∅12(ADD)``

**Zone length** (dimension text near stirrup annotation):
    ``140``  or  ``458``  (numeric dimension value)

The diameter symbol can be any of: ``%%c  %%C  ∅  Ø  ⌀  ø  φ  Φ  T``

Zone types
----------
- **start** : first stirrup zone (left end of beam)
- **mid**   : middle stirrup zone(s)
- **end**   : last stirrup zone (right end of beam)

ADD stirrups are counted only for start/end zones, not for mid zones.
"""

from __future__ import annotations

import math
import os
import re
import shutil
import time
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# ── Fix pywin32 DLL loading ────────────────────────────────────────
_pywin32_system32 = os.path.join(
    os.path.dirname(os.path.dirname(os.__file__)),
    "Lib", "site-packages", "pywin32_system32",
)
if os.path.isdir(_pywin32_system32):
    os.add_dll_directory(_pywin32_system32)
    if _pywin32_system32 not in os.environ.get("PATH", ""):
        os.environ["PATH"] = _pywin32_system32 + os.pathsep + os.environ.get("PATH", "")

import pythoncom        # noqa: E402,I001
import win32com.client  # noqa: E402

from civiltools.building.listofer_style import (  # noqa: E402
    CELL_HEIGHT as DEFAULT_STIRRUP_TABLE_CELL_H,
    DEFAULT_LISTOFER_TEMPLATE,
    DESCRIPTION_COL_WIDTH,
    DIA_COL_WIDTH as DEFAULT_STIRRUP_TABLE_DIA_COL_WIDTH,
    HEADER_FILL_COLOR,
    MIN_TEXT_HEIGHT as DEFAULT_STIRRUP_TABLE_MIN_TEXT_H,
    POS_COL_WIDTH,
    ROW_COL_WIDTH,
    SHAPE_COL_WIDTH,
    SUMMARY_FILL_COLOR,
    TEXT_HEIGHT as DEFAULT_STIRRUP_TABLE_TEXT_H,
    TEXT_HEIGHT_FACTOR as DEFAULT_STIRRUP_TABLE_TEXT_FACTOR,
    UNIT_WEIGHT_COL_WIDTH,
    resolve_text_height,
)

# ---------------------------------------------------------------------------
#  Constants
# ---------------------------------------------------------------------------

STEEL_DENSITY = 7850.0          # kg/m³
CONCRETE_COVER_CM = 4.0         # concrete cover for beams (cm)
DEFAULT_HOOK_FACTOR = 10.0      # hook length = factor × diameter
STANDARD_BAR_LENGTH_M = 12.0    # standard rebar stock length
# Fixed columns: Row, POS, Description, Shape (shared widths) + Zone, Spc,
# ZoneLen, Cnt, B, H, SingleL, TotalL, UnitW (stirrup-specific/shared widths)
DEFAULT_STIRRUP_TABLE_COL_WIDTHS = [
    ROW_COL_WIDTH, POS_COL_WIDTH, DESCRIPTION_COL_WIDTH, SHAPE_COL_WIDTH,
    12, 10, 14, 8, 8, 8, 14, 12, UNIT_WEIGHT_COL_WIDTH,
]

# All AutoCAD representations of the diameter symbol
_DIA = r'(?:%%[cC]|[∅Ø⌀øφΦ~T])'

# ---------------------------------------------------------------------------
#  Regex patterns
# ---------------------------------------------------------------------------

# Stirrup spacing: T12@15 or ∅12@15
RE_STIRRUP_SPACING = re.compile(
    rf'\s*{_DIA}\s*(\d+)\s*@\s*(\d+(?:\.\d+)?)\s*',
    re.IGNORECASE,
)

# Stirrup ADD: 1T12(ADD) or 1∅12(ADD)
RE_STIRRUP_ADD = re.compile(
    rf'(\d+)\s*{_DIA}\s*(\d+)\s*\(\s*ADD\s*\)',
    re.IGNORECASE,
)

# Zone length dimension: standalone number like 140, 458
# Must be a standalone number, NOT part of an axis label (L14, P19, etc.)
# Excludes numbers preceded by letters (axis labels) or followed by letters
RE_ZONE_LENGTH = re.compile(
    r'(?:^|[^A-Za-z\d.])(\d{2,4})(?:\s*(?:cm|CM))?\s*$',
    re.IGNORECASE,
)

# Fallback: extract any 2-4 digit number from text
# But exclude axis labels like L14, P19, A1, B2
RE_ANY_NUMBER = re.compile(
    r'(?<![A-Za-z])(\d{2,4})(?![A-Za-z])',
)

# Beam size: 60x70 or 60×70
RE_BEAM_SIZE = re.compile(
    r'(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)',
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
#  Data classes
# ---------------------------------------------------------------------------

@dataclass
class StirrupTextInfo:
    """One Text / MText / Dimension entity read from AutoCAD."""

    obj_id: int
    text_string: str
    insertion_point: tuple[float, float, float]
    rotation: float = 0.0
    text_height: float = 1.0
    entity_type: str = "TEXT"  # 'TEXT', 'MTEXT', 'DIMENSION'
    measurement: float | None = None  # DIMENSION numeric value when available
    attr_tag: str | None = None  # block attribute tag (e.g. PO, Des1)


@dataclass
class StirrupZone:
    """One stirrup zone (a region with uniform spacing)."""

    pos: int = 0                    # POS number (S01, S02, ...)
    zone_type: str = ""             # 'start', 'mid', 'end'
    diameter: float = 0.0           # mm
    spacing: float = 0.0            # cm
    zone_length: float = 0.0      # cm
    count: int = 0                  # total stirrup count
    single_length: float = 0.0      # m (length of one stirrup)
    total_length: float = 0.0     # m
    unit_weight: float = 0.0      # kg/m
    total_weight: float = 0.0     # kg
    has_add: bool = False         # has ADD stirrup?
    add_count: int = 0            # number of ADD stirrups
    description: str = ""           # e.g. "T12@15 + ADD"
    beam: BeamDimensions = field(default_factory=lambda: BeamDimensions())
    text_ids: list[int] = field(default_factory=list)
    raw_texts: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def __post_init__(self):
        if not self.description:
            add_text = " + ADD" if self.has_add else ""
            self.description = f"T{int(self.diameter)}@{int(self.spacing)}{add_text}"

    def check_completeness(self) -> bool:
        """Validate that all required fields are present."""
        existing = set(self.errors)
        self.errors = list(existing)
        if self.diameter <= 0:
            existing.add("diameter")
        if self.spacing <= 0:
            existing.add("spacing")
        if self.zone_length <= 0:
            existing.add("zone_length")

        self.errors = sorted(existing)
        return len(self.errors) == 0


@dataclass
class BeamDimensions:
    """Beam cross-section dimensions."""

    width: float = 60.0     # cm
    height: float = 70.0    # cm


# ---------------------------------------------------------------------------
#  Helper functions
# ---------------------------------------------------------------------------

def _spoint(x: float, y: float):
    """Create an AutoCAD 2-D point VARIANT."""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y))


def _dist(p1: tuple, p2: tuple) -> float:
    """2-D Euclidean distance between two points."""
    return math.hypot(p1[0] - p2[0], p1[1] - p2[1])


def calculate_stirrup_length(
    diameter_mm: float,
    beam: BeamDimensions,
    cover_cm: float = CONCRETE_COVER_CM,
    hook_factor: float = DEFAULT_HOOK_FACTOR,
) -> float:
    """Calculate the cutting length of one closed stirrup (metres).

    Formula::

        U = 2×(h_eff) + 2×(b_eff) + 2×hook

    where::

        h_eff = beam.height − 2×cover
        b_eff = beam.width  − 2×cover
        hook  = hook_factor × d_cm

    Args:
        diameter_mm: stirrup bar diameter in millimetres.
        beam: BeamDimensions instance.
        cover_cm: concrete cover in centimetres.
        hook_factor: hook length = factor × bar diameter.

    Returns:
        Cutting length in metres, rounded to 3 decimals.
    """
    d_cm = diameter_mm / 10.0
    h_eff = beam.height - 2 * cover_cm
    b_eff = beam.width - 2 * cover_cm
    hook_length = hook_factor * d_cm / 100.0  # cm → m

    perimeter = 2 * (h_eff + b_eff) / 100.0   # cm → m
    total = perimeter + 2 * hook_length
    return round(total, 3)


def calculate_unit_weight(diameter_mm: float) -> float:
    """Return the unit weight of a rebar (kg/m).

    Formula: ρ × π × (d/2)²
    """
    d_m = diameter_mm / 1000.0
    area = math.pi * (d_m / 2) ** 2
    return round(STEEL_DENSITY * area, 3)


def calculate_stirrup_count(
    zone_length_cm: float,
    spacing_cm: float,
    has_add: bool = False,
) -> int:
    """Number of stirrups in a zone.

    Formula::

        count = round(zone_length / spacing) + 1

    If *has_add* is True (start/end zone), add one extra stirrup.
    """
    base = round(zone_length_cm / spacing_cm) + 1
    return base + (1 if has_add else 0)


# ---------------------------------------------------------------------------
#  Main engine
# ---------------------------------------------------------------------------

class StirrupFromDwg:
    """Read and parse stirrup annotation text from an open AutoCAD drawing."""

    PROXIMITY_FACTOR: int = 25  # max search distance = text_height × factor

    def __init__(
        self,
        beam_dims: BeamDimensions | None = None,
        cover_cm: float = CONCRETE_COVER_CM,
        hook_factor: float = DEFAULT_HOOK_FACTOR,
    ) -> None:
        self.acad: Any = win32com.client.Dispatch("AutoCAD.Application")
        self.acad.Visible = True
        self.doc: Any = self.acad.ActiveDocument
        self.text_objects: list[StirrupTextInfo] = []
        self.stirrup_zones: list[StirrupZone] = []
        self.beam = beam_dims or BeamDimensions()
        self.cover = cover_cm
        self.hook_factor = hook_factor

    # ------------------------------------------------------------------
    #  MText cleanup
    # ------------------------------------------------------------------
    @staticmethod
    def clean_mtext(text: str) -> str:
        """Strip MText formatting codes, returning plain text."""
        text = re.sub(r'{\[^;]+;([^}]*)}', r'', text)
        text = re.sub(r'\[AHWQTLOoPpCcFf][^;]*;', '', text)
        text = text.replace('\\P', ' ').replace('\\p', ' ')
        text = text.replace('{', '').replace('}', '')
        return ' '.join(text.split()).strip()

    # ------------------------------------------------------------------
    #  Read text entities
    # ------------------------------------------------------------------
    def _read_text_obj(self, obj: Any) -> StirrupTextInfo | None:
        """Convert one COM text entity to a *StirrupTextInfo*, or *None*."""
        if obj.ObjectName not in ("AcDbText", "AcDbMText"):
            return None
        raw = obj.TextString
        text = self.clean_mtext(raw) if obj.ObjectName == "AcDbMText" else raw
        ip = obj.InsertionPoint
        return StirrupTextInfo(
            obj_id=obj.ObjectID,
            text_string=text,
            insertion_point=(ip[0], ip[1], ip[2] if len(ip) > 2 else 0.0),
            rotation=getattr(obj, 'Rotation', 0),
            text_height=getattr(obj, 'Height', 1.0),
            entity_type='MTEXT' if obj.ObjectName == "AcDbMText" else 'TEXT',
        )

    def _read_dim_obj(self, dim: Any) -> StirrupTextInfo | None:
        """Read a DIMENSION entity and extract its measurement value."""
        try:
            # Get dimension text (may be overridden or measured)
            dim_text = ""
            measurement = 0.0

            # Try to get measurement value
            if hasattr(dim, 'Measurement'):
                measurement = float(dim.Measurement)

            # Try to get text string
            if hasattr(dim, 'TextString'):
                dim_text = str(dim.TextString)

            # For these stirrup drawings, the displayed dimension text is the
            # value we want to use as zone length (e.g. 140, 458). Keep the
            # raw measurement as a fallback only.
            text_value = None
            if dim_text and dim_text.strip() not in ("", "<>", "<", ">"):
                m = RE_ANY_NUMBER.search(dim_text)
                if m:
                    text_value = float(m.group(1))

            # If text is empty or "<>", use the measurement value
            if text_value is None:
                if measurement > 0:
                    text_value = float(int(round(measurement)))
                    dim_text = str(int(round(measurement)))
                else:
                    return None
            else:
                # Normalize the text so parse_stirrups can consume it directly.
                dim_text = str(int(round(text_value)))

            # Get text position
            ip = (0.0, 0.0, 0.0)
            if hasattr(dim, 'TextPosition'):
                tp = dim.TextPosition
                ip = (tp[0], tp[1], tp[2] if len(tp) > 2 else 0.0)
            elif hasattr(dim, 'InsertionPoint'):
                ip = dim.InsertionPoint

            # Get text height
            th = 2.5
            if hasattr(dim, 'TextHeight'):
                th = float(dim.TextHeight)

            return StirrupTextInfo(
                obj_id=dim.ObjectID,
                text_string=dim_text,
                insertion_point=ip,
                rotation=getattr(dim, 'Rotation', 0),
                text_height=th,
                entity_type='DIMENSION',
                measurement=measurement if measurement > 0 else None,
            )
        except Exception:
            return None

    def _read_block_attributes(self, block_ref: Any) -> list[StirrupTextInfo]:
        """Read visible attributes from a BLOCK REFERENCE as text-like inputs."""
        out: list[StirrupTextInfo] = []
        try:
            attrs = block_ref.GetAttributes()
        except Exception:
            attrs = []

        # Fallback point if attribute insertion point is unavailable
        blk_ip = (0.0, 0.0, 0.0)
        try:
            ip = block_ref.InsertionPoint
            blk_ip = (ip[0], ip[1], ip[2] if len(ip) > 2 else 0.0)
        except Exception:
            pass

        for at in attrs:
            try:
                txt = str(getattr(at, 'TextString', '') or '').strip()
                if not txt:
                    continue

                at_ip = blk_ip
                try:
                    p = at.InsertionPoint
                    at_ip = (p[0], p[1], p[2] if len(p) > 2 else 0.0)
                except Exception:
                    pass

                out.append(
                    StirrupTextInfo(
                        obj_id=at.ObjectID,
                        text_string=txt,
                        insertion_point=at_ip,
                        rotation=getattr(at, 'Rotation', 0),
                        text_height=getattr(at, 'Height', 1.0),
                        entity_type='ATTRIB',
                        attr_tag=str(getattr(at, 'TagString', '') or '').strip(),
                    )
                )
            except Exception:
                continue

        return out

    def get_text_objects_from_selection(self) -> list[StirrupTextInfo]:
        """Prompt the user to select objects; keep Text/MText/Dimension/Block-Attributes."""
        texts: list[StirrupTextInfo] = []
        ss_name = f"_StirrupSel_{int(time.time())}"

        try:
            # Make sure AutoCAD is active
            self.acad.Visible = True
            try:
                self.acad.WindowState = 1  # acMax
            except Exception:
                pass
            time.sleep(0.5)

            ss_col = self.doc.SelectionSets
            # Clean up old selection set
            for i in range(ss_col.Count):
                try:
                    if ss_col.Item(i).Name == ss_name:
                        ss_col.Item(i).Delete()
                        break
                except Exception:
                    pass

            ss = ss_col.Add(ss_name)

            try:
                self.acad.Activate()
            except Exception:
                pass

            print("Select stirrup texts in AutoCAD and press Enter …")
            print("  (Make sure AutoCAD window is active)")

            try:
                ss.SelectOnScreen()
            except Exception as exc:
                print(f"SelectOnScreen failed: {exc}")
                print("Falling back to selecting all text objects...")
                ss.Delete()
                return self.get_all_text_objects()

            for i in range(ss.Count):
                obj = ss.Item(i)
                obj_name = obj.ObjectName

                if obj_name in ("AcDbText", "AcDbMText"):
                    info = self._read_text_obj(obj)
                    if info:
                        texts.append(info)
                elif obj_name in ("AcDbDimension", "AcDbAlignedDimension"):
                    info = self._read_dim_obj(obj)
                    if info:
                        texts.append(info)
                elif obj_name == "AcDbBlockReference":
                    texts.extend(self._read_block_attributes(obj))

            ss.Delete()

        except Exception as exc:
            print(f"Selection error: {exc}")
            try:
                self.doc.SelectionSets.Item(ss_name).Delete()
            except Exception:
                pass

        self.text_objects = texts
        print(f"Selected {len(texts)} text/dimension/attribute objects")
        return texts

    def get_all_text_objects(self) -> list[StirrupTextInfo]:
        """Get ALL Text, MText, Dimension & block attributes from model space."""
        all_texts: list[StirrupTextInfo] = []

        # Iterate directly through modelspace instead of using Select with filters
        # This avoids filter compatibility issues across AutoCAD versions
        try:
            msp = self.doc.ModelSpace
            for i in range(msp.Count):
                obj = msp.Item(i)
                obj_name = obj.ObjectName

                if obj_name in ("AcDbText", "AcDbMText"):
                    info = self._read_text_obj(obj)
                    if info:
                        all_texts.append(info)
                elif obj_name in ("AcDbDimension", "AcDbAlignedDimension"):
                    info = self._read_dim_obj(obj)
                    if info:
                        all_texts.append(info)
                elif obj_name == "AcDbBlockReference":
                    all_texts.extend(self._read_block_attributes(obj))

        except Exception as exc:
            print(f"ModelSpace iteration error: {exc}")

        self.text_objects = all_texts
        print(f"Found {len(all_texts)} text/dimension/attribute objects")
        return all_texts

    # ------------------------------------------------------------------
    #  Parse → StirrupZone
    # ------------------------------------------------------------------
    def parse_stirrups(self) -> list[StirrupZone]:
        """Classify collected text objects into StirrupZone entries."""
        self.stirrup_zones = []
        stirrup_texts: list[tuple[StirrupTextInfo, float, float]] = []
        add_texts: list[tuple[StirrupTextInfo, int, float]] = []
        length_texts: list[tuple[StirrupTextInfo, float]] = []
        beam_size_texts: list[tuple[StirrupTextInfo, float, float]] = []
        used_ids: set[int] = set()

        print(f"\n{'='*50}")
        print(f"PARSING {len(self.text_objects)} TEXT OBJECTS")
        print(f"{'='*50}")

        for i, ti in enumerate(self.text_objects):
            raw = ti.text_string
            print(
                f"  [{i+1}] type={ti.entity_type:12s} | text='{raw[:40]:40s}' | "
                f"pos=({ti.insertion_point[0]:.1f}, {ti.insertion_point[1]:.1f})"
            )

        for ti in self.text_objects:
            txt = ti.text_string.strip()

            # Skip empty texts
            if not txt:
                continue

            # Ignore helper POS/PO attribute values (block bookkeeping),
            # otherwise they can be misread as zone lengths (e.g. "73").
            if ti.entity_type == 'ATTRIB' and (ti.attr_tag or '').upper() in {'PO', 'POS'}:
                print(f"    → SKIPPED (attribute tag={ti.attr_tag})")
                continue

            # 1) Stirrup spacing text: T12@15
            m = RE_STIRRUP_SPACING.search(txt)
            if m:
                stirrup_texts.append(
                    (ti, float(m.group(1)), float(m.group(2))))
                used_ids.add(ti.obj_id)
                print(f"    → STIRRUP: d={m.group(1)}mm @ {m.group(2)}cm")
                continue

            # 2) Stirrup ADD text: 1T12(ADD)
            m = RE_STIRRUP_ADD.search(txt)
            if m:
                add_texts.append(
                    (ti, int(m.group(1)), float(m.group(2))))
                used_ids.add(ti.obj_id)
                print(f"    → ADD: {m.group(1)}x{m.group(2)}mm")
                continue

            # 3) Zone length text: 140, 458 (numeric) - from
            # Dimension entity - use measurement directly
            if ti.entity_type == 'DIMENSION':
                m = RE_ZONE_LENGTH.search(txt)
                if m:
                    val = float(m.group(1))
                    if 20 <= val <= 1000:  # reasonable zone length range
                        length_texts.append((ti, val))
                        used_ids.add(ti.obj_id)
                        print(f"    → LENGTH: {val}cm")
                        continue
                    else:
                        print(f"    → REJECTED length: {val}cm (out of range)")


            # 4) Beam size: 60x70
            m = RE_BEAM_SIZE.search(txt)
            if m:
                beam_size_texts.append((ti, float(m.group(1)), float(m.group(2))))
                used_ids.add(ti.obj_id)
                print(f"    → BEAM SIZE: {m.group(1)}x{m.group(2)}")
                continue

            # 6) Not a stirrup text → skip
            print("    → SKIPPED")

        # ---- Match stirrup texts with nearest length text ----
        print(f"\n{'-'*50}")
        print("CLASSIFICATION RESULTS:")
        print(f"  Stirrup texts: {len(stirrup_texts)}")
        print(f"  Length texts:  {len(length_texts)}")
        print(f"  ADD texts:     {len(add_texts)}")
        print(f"  Beam sizes:    {len(beam_size_texts)}")
        print(f"{'-'*50}")

        for ti, dia, spc in stirrup_texts:
            print(f"  STIRRUP: d={dia}mm @ {spc}cm | '{ti.text_string}'")
        for li, lv in length_texts:
            print(f"  LENGTH:  {lv}cm | '{li.text_string}'")
        for ai, cnt, dia in add_texts:
            print(f"  ADD:     {cnt}x{dia}mm | '{ai.text_string}'")
        for bi, bw, bh in beam_size_texts:
            print(f"  BEAM:    {bw}x{bh}cm | '{bi.text_string}'")

        used_length_ids: set[int] = set()
        for ti, dia, spc in stirrup_texts:
            zone = StirrupZone(
                diameter=dia,
                spacing=spc,
                text_ids=[ti.obj_id],
                raw_texts=[ti.text_string],
            )
            best_dist = float('inf')
            best_idx = -1
            max_d = ti.text_height * self.PROXIMITY_FACTOR

            for idx, (li, lv) in enumerate(length_texts):
                if li.obj_id in used_length_ids:
                    continue
                d = _dist(ti.insertion_point, li.insertion_point)
                if d < best_dist and d <= max_d:
                    best_dist = d
                    best_idx = idx

            if best_idx >= 0:
                li, lv = length_texts[best_idx]
                zone.zone_length = lv
                zone.text_ids.append(li.obj_id)
                zone.raw_texts.append(li.text_string)
                used_length_ids.add(li.obj_id)

            # Find nearest beam size for THIS stirrup
            best_beam_dist = float('inf')
            best_beam = None
            max_beam_d = ti.text_height * self.PROXIMITY_FACTOR * 3  # larger radius for beam size

            for bti, bw, bh in beam_size_texts:
                d = _dist(ti.insertion_point, bti.insertion_point)
                if d < best_beam_dist and d <= max_beam_d:
                    best_beam_dist = d
                    best_beam = BeamDimensions(width=bw, height=bh)

            if best_beam:
                zone.beam = best_beam
                print(
                    "    → BEAM MATCHED: "
                    f"{best_beam.width}x{best_beam.height}cm "
                    f"(dist={best_beam_dist:.1f})"
                )
            else:
                zone.beam = BeamDimensions(width=self.beam.width, height=self.beam.height)
                print(
                    "    → BEAM FALLBACK: "
                    f"{zone.beam.width}x{zone.beam.height}cm "
                    "(no nearby size found)"
                )

            # Check for ADD near this stirrup text
            zone.has_add, zone.add_count = self._find_add_nearby(
                ti.insertion_point, add_texts, dia)

            # Calculate derived values
            # single_length is always calculated (doesn't depend on zone_length)
            zone.single_length = calculate_stirrup_length(
                zone.diameter, zone.beam, self.cover, self.hook_factor)
            zone.unit_weight = calculate_unit_weight(zone.diameter)

            if zone.zone_length > 0:
                zone.count = calculate_stirrup_count(
                    zone.zone_length, zone.spacing, zone.has_add)
                zone.total_length = round(zone.count * zone.single_length, 2)
                zone.total_weight = round(zone.total_length * zone.unit_weight, 2)
            else:
                # No zone length found — flag as incomplete
                zone.errors.append("zone_length_not_found")

            zone.check_completeness()
            if "zone_length_not_found" in zone.errors and zone.zone_length > 0:
                zone.errors.remove("zone_length_not_found")
            self.stirrup_zones.append(zone)
            print(f"    → ZONE CREATED: POS=S{zone.pos:02d}, type={zone.zone_type}, "
                  f"d={zone.diameter}mm, spc={zone.spacing}cm, len={zone.zone_length}cm, "
                  f"count={zone.count}, single={zone.single_length}m, "
                  f"total={zone.total_length}m, weight={zone.total_weight}kg, "
                  f"beam={zone.beam.width}x{zone.beam.height}")
            if zone.errors:
                print(f"       ERRORS: {zone.errors}")

        # ---- Sort by X position (left to right) ----
        self.stirrup_zones.sort(
            key=lambda z: min(
                (t.insertion_point[0] for t in self.text_objects
                 if t.obj_id in z.text_ids),
                default=0.0
            )
        )

        # ---- Assign POS numbers and zone types ----
        self._assign_pos_and_types()

        return self.stirrup_zones

    def _find_add_nearby(
        self,
        stirrup_pos: tuple,
        add_texts: list[tuple[StirrupTextInfo, int, float]],
        diameter: float,
        search_radius: float = 30.0,
    ) -> tuple[bool, int]:
        """Check if an ADD text exists near the stirrup annotation."""
        for ti, count, dia in add_texts:
            if dia != diameter:
                continue
            d = _dist(stirrup_pos, ti.insertion_point)
            if d <= search_radius:
                return True, count
        return False, 0

    def _assign_pos_and_types(self):
        """Assign POS numbers and zone types (start/mid/end)."""
        n = len(self.stirrup_zones)
        for i, zone in enumerate(self.stirrup_zones, 1):
            zone.pos = i
            if n == 1:
                zone.zone_type = 'mid'
            elif i == 1:
                zone.zone_type = 'start'
            elif i == n:
                zone.zone_type = 'end'
            else:
                zone.zone_type = 'mid'

    # ------------------------------------------------------------------
    #  Post-processing helpers
    # ------------------------------------------------------------------
    def get_incomplete_ids(self) -> list[int]:
        """ObjectIDs of text entities belonging to incomplete stirrup zones."""
        ids: list[int] = []
        for zone in self.stirrup_zones:
            if zone.errors:
                ids.extend(zone.text_ids)
        return ids

    def highlight_incomplete(self) -> int:
        """Highlight incomplete stirrup texts and zoom to them in AutoCAD.

        Returns the number of highlighted items.
        """
        ids = self.get_incomplete_ids()
        if not ids:
            return 0

        min_x = min_y = float('inf')
        max_x = max_y = float('-inf')

        for obj_id in ids:
            try:
                obj = self.doc.ObjectIdToObject(obj_id)
                obj.Highlight(True)
                mn, mx = obj.GetBoundingBox()
                min_x = min(min_x, mn[0])
                min_y = min(min_y, mn[1])
                max_x = max(max_x, mx[0])
                max_y = max(max_y, mx[1])
            except Exception:
                pass

        if min_x < float('inf'):
            dx = max(max_x - min_x, 1) * 0.15
            dy = max(max_y - min_y, 1) * 0.15
            try:
                self.acad.ZoomWindow(
                    _spoint(min_x - dx, min_y - dy),
                    _spoint(max_x + dx, max_y + dy),
                )
            except Exception:
                try:
                    self.doc.SendCommand("ZOOM E\n")
                except Exception:
                    pass

        return len(ids)

    def summary(self) -> dict[str, int]:
        """Quick summary of parsed stirrup zones."""
        total = len(self.stirrup_zones)
        ok = sum(1 for z in self.stirrup_zones if not z.errors)
        return {
            'total': total,
            'complete': ok,
            'incomplete': total - ok,
            'start_zones': sum(1 for z in self.stirrup_zones if z.zone_type == 'start'),
            'mid_zones': sum(1 for z in self.stirrup_zones if z.zone_type == 'mid'),
            'end_zones': sum(1 for z in self.stirrup_zones if z.zone_type == 'end'),
        }

    def summary_by_size(self) -> list[dict[str, Any]]:
        """Group complete stirrups by diameter and return summary rows.

        Each row: size (mm), total_length (m), number (ceil of
        total_length / 12 m), weight (kg), unit_weight (kg/m).
        """
        groups: dict[float, float] = defaultdict(float)
        for zone in self.stirrup_zones:
            if zone.diameter <= 0 or zone.total_length <= 0:
                continue
            groups[zone.diameter] += zone.total_length

        rows: list[dict[str, Any]] = []
        grand_length = 0.0
        grand_number = 0
        grand_weight = 0.0

        for dia in sorted(groups):
            total_l = groups[dia]
            unit_w = calculate_unit_weight(dia)
            weight = round(total_l * unit_w, 2)
            n_bars = math.ceil(total_l / STANDARD_BAR_LENGTH_M)

            grand_length += total_l
            grand_number += n_bars
            grand_weight += weight

            rows.append({
                'Size (mm)': int(dia),
                'Total Length (m)': round(total_l, 2),
                'Number (12 m bars)': n_bars,
                'Unit Weight (kg/m)': unit_w,
                'Total Weight (kg)': weight,
            })

        if rows:
            rows.append({
                'Size (mm)': 'TOTAL',
                'Total Length (m)': round(grand_length, 2),
                'Number (12 m bars)': grand_number,
                'Unit Weight (kg/m)': '',
                'Total Weight (kg)': round(grand_weight, 2),
            })

        return rows


# ---------------------------------------------------------------------------
#  Excel Export
# ---------------------------------------------------------------------------

def export_to_excel(
    extractor: StirrupFromDwg,
    output_path: str,
    project_name: str = "Project",
) -> str:
    """Export stirrup schedule to an Excel file.

    Columns: Row | POS | Description | Zone Type | Diameter | Spacing |
             Zone Length | Count | Single Length | Total Length |
             Unit Weight | Total Weight
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Install: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create the Stirrups worksheet")
    ws.title = "Stirrups"
    ws.sheet_view.rightToLeft = True

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Header
    headers = [
        "Row", "POS", "Description", "Zone Type", "Dia (mm)", "Spacing (cm)",
        "Zone Len (cm)", "Count", "Single Len (m)", "Total Len (m)",
        "Unit Wt (kg/m)", "Total Wt (kg)",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    zone_type_map = {'start': 'Start', 'mid': 'Mid', 'end': 'End'}
    row = 2
    for i, zone in enumerate(extractor.stirrup_zones, 1):
        data = [
            i,
            f"S{zone.pos:02d}",
            zone.description,
            zone_type_map.get(zone.zone_type, zone.zone_type),
            int(zone.diameter) if zone.diameter else "",
            int(zone.spacing) if zone.spacing else "",
            int(zone.zone_length) if zone.zone_length else "",
            zone.count,
            zone.single_length,
            zone.total_length,
            zone.unit_weight,
            zone.total_weight,
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        row += 1

    # Summary row
    summary = extractor.summary_by_size()
    if summary:
        ws.cell(row=row, column=1, value="").border = thin_border
        ws.cell(row=row, column=2, value="").border = thin_border
        ws.cell(row=row, column=3, value="Grand Total").font = Font(bold=True)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3).border = thin_border

        total = summary[-1]
        for col in range(4, 8):
            ws.cell(row=row, column=col, value="").border = thin_border
        ws.cell(row=row, column=8, value=total.get('Number (12 m bars)', '')).font = Font(bold=True)
        ws.cell(row=row, column=8).border = thin_border
        ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=9, value="").border = thin_border
        ws.cell(row=row, column=10, value=total.get('Total Length (m)', '')).font = Font(bold=True)
        ws.cell(row=row, column=10).border = thin_border
        ws.cell(row=row, column=10).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=11, value="").border = thin_border
        ws.cell(row=row, column=12, value=total.get('Total Weight (kg)', '')).font = Font(bold=True)
        ws.cell(row=row, column=12).border = thin_border
        ws.cell(row=row, column=12).alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = [6, 8, 20, 10, 10, 12, 14, 8, 14, 12, 14, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
#  AutoCAD Table Drawer
# ---------------------------------------------------------------------------

class StirrupTableDrawer:
    """Draw a stirrup schedule table directly in AutoCAD."""

    def __init__(
        self,
        doc: Any = None,
        template_path: str | Path | None = DEFAULT_LISTOFER_TEMPLATE,
        block_scale: float = 0.08,
    ) -> None:
        if doc is None:
            self.acad = win32com.client.Dispatch("AutoCAD.Application")
            self.doc = self.acad.ActiveDocument
        else:
            self.doc = doc
        self.msp = self.doc.ModelSpace
        self.template_path = Path(template_path) if template_path else None
        self.block_scale = float(block_scale)
        self._template_blocks_ready = False

    def _get_ezdxf(self):
        try:
            import ezdxf
        except ImportError as exc:
            raise ImportError(
                "Install ezdxf to load listofer template blocks (pip install ezdxf)."
            ) from exc
        return ezdxf

    def _entity_color(self, ent: Any) -> int | None:
        try:
            color = int(ent.dxf.color)
            return color if color > 0 else None
        except Exception:
            return None

    def _copy_template_entity_to_block(self, blk_com: Any, ent: Any) -> None:
        etype = ent.dxftype()
        color = self._entity_color(ent)
        created = None

        if etype == "LINE":
            start = ent.dxf.start
            end = ent.dxf.end
            created = blk_com.AddLine(
                win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    (start.x, start.y, getattr(start, "z", 0.0)),
                ),
                win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    (end.x, end.y, getattr(end, "z", 0.0)),
                ),
            )
        elif etype == "ARC":
            center = ent.dxf.center
            created = blk_com.AddArc(
                win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    (center.x, center.y, getattr(center, "z", 0.0)),
                ),
                float(ent.dxf.radius),
                float(ent.dxf.start_angle) * math.pi / 180.0,
                float(ent.dxf.end_angle) * math.pi / 180.0,
            )
        elif etype == "CIRCLE":
            center = ent.dxf.center
            created = blk_com.AddCircle(
                win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    (center.x, center.y, getattr(center, "z", 0.0)),
                ),
                float(ent.dxf.radius),
            )
        elif etype == "LWPOLYLINE":
            points = list(ent.get_points("xyb"))
            if len(points) >= 2:
                flat_points: list[float] = []
                for px, py, _bulge in points:
                    flat_points.extend([px, py])
                polyline = blk_com.AddLightWeightPolyline(
                    win32com.client.VARIANT(
                        pythoncom.VT_ARRAY | pythoncom.VT_R8,
                        tuple(flat_points),
                    )
                )
                for index, (_px, _py, bulge) in enumerate(points):
                    if abs(float(bulge or 0.0)) > 1e-9:
                        try:
                            polyline.SetBulge(index, float(bulge))
                        except Exception:
                            pass
                try:
                    polyline.Closed = bool(ent.closed)
                except Exception:
                    pass
                created = polyline
        elif etype == "TEXT":
            insert = ent.dxf.insert
            created = blk_com.AddText(
                str(ent.dxf.text),
                win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    (insert.x, insert.y, getattr(insert, "z", 0.0)),
                ),
                float(getattr(ent.dxf, "height", 1.0) or 1.0),
            )
            try:
                created.Rotation = float(getattr(ent.dxf, "rotation", 0.0) or 0.0) * math.pi / 180.0
            except Exception:
                pass
        elif etype == "ATTDEF":
            insert = ent.dxf.insert
            created = blk_com.AddAttribute(
                float(getattr(ent.dxf, "height", 1.0) or 1.0),
                0,
                str(getattr(ent.dxf, "prompt", ent.dxf.tag)),
                win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    (insert.x, insert.y, getattr(insert, "z", 0.0)),
                ),
                str(ent.dxf.tag),
                str(getattr(ent.dxf, "text", "") or ""),
            )

        if created is not None and color is not None:
            try:
                created.Color = color
            except Exception:
                pass

    def _ensure_template_blocks(self) -> None:
        if self._template_blocks_ready:
            return
        if self.template_path is None or not self.template_path.exists():
            return

        try:
            blocks = self.doc.Blocks
            existing = {blocks.Item(i).Name for i in range(blocks.Count)}
        except Exception:
            existing = set()

        if "TO" in existing:
            self._template_blocks_ready = True
            return

        ezdxf = self._get_ezdxf()
        readfile_fn = getattr(ezdxf, "readfile", None)
        if readfile_fn is None:
            from ezdxf import filemanagement as _filemanagement

            readfile_fn = _filemanagement.readfile
        template_doc = readfile_fn(str(self.template_path))
        if "TO" not in template_doc.blocks:
            return

        try:
            block_def = self.doc.Blocks.Add(
                win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0)),
                "TO",
            )
        except Exception:
            return

        for ent in template_doc.blocks.get("TO"):
            try:
                self._copy_template_entity_to_block(block_def, ent)
            except Exception:
                continue

        self._template_blocks_ready = True

    def _set_block_attributes(self, block_ref: Any, values: dict[str, str]) -> None:
        try:
            attributes = block_ref.GetAttributes()
        except Exception:
            return
        for attrib in attributes:
            try:
                tag = str(getattr(attrib, "TagString", "") or "").strip().upper()
                if tag in values:
                    attrib.TextString = str(values[tag])
            except Exception:
                continue

    def _stirrup_shape_values(
        self,
        zone: StirrupZone,
        cover_cm: float,
        hook_factor: float,
    ) -> dict[str, str]:
        width_cm = max(float(zone.beam.width) - 2.0 * cover_cm, 0.0)
        height_cm = max(float(zone.beam.height) - 2.0 * cover_cm, 0.0)
        hook_cm = max(hook_factor * (float(zone.diameter) / 10.0), 0.0)
        return {
            "L1": str(int(round(width_cm))),
            "L2": str(int(round(height_cm))),
            "L3": str(int(round(hook_cm))),
        }

    def _insert_shape_block(
        self,
        zone: StirrupZone,
        x: float,
        y_top: float,
        width: float,
        height: float,
        cover_cm: float,
        hook_factor: float,
    ) -> None:
        self._ensure_template_blocks()
        insert_x = x + width * 0.5
        insert_y = y_top - height * 0.60
        try:
            block_ref = self.msp.InsertBlock(
                win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    (insert_x, insert_y, 0.0),
                ),
                "TO",
                self.block_scale,
                self.block_scale,
                self.block_scale,
                0.0,
            )
        except Exception as exc:
            print(f"Stirrup shape block insertion failed for S{zone.pos:02d}: {exc}")
            return

        self._set_block_attributes(
            block_ref,
            self._stirrup_shape_values(zone, cover_cm, hook_factor),
        )

    @staticmethod
    def _current_doc_dir(doc: Any) -> Path:
        """Return directory of active AutoCAD document, fallback to cwd."""
        try:
            full_name = str(getattr(doc, "FullName", "") or "").strip()
            if full_name:
                return Path(full_name).resolve().parent
        except Exception:
            pass
        return Path.cwd().resolve()

    def _resolve_output_dxf_path(
        self,
        output_path: str | Path | None,
        filename: str | None,
        prefix: str = "stirrup_listofer",
    ) -> Path:
        """Build output DXF path next to current DWG unless explicit path is passed."""
        if output_path is not None:
            return Path(output_path).resolve()

        out_dir = self._current_doc_dir(self.doc)
        if filename:
            name = filename
            if not name.lower().endswith(".dxf"):
                name += ".dxf"
        else:
            name = f"{prefix}_{uuid.uuid4().hex[:8]}.dxf"
        return out_dir / name

    @staticmethod
    def _draw_cell_dxf(msp: Any, x: float, y_top: float, width: float, height: float) -> None:
        """Draw one rectangular cell in a DXF modelspace."""
        pts = [
            (x, y_top),
            (x + width, y_top),
            (x + width, y_top - height),
            (x, y_top - height),
            (x, y_top),
        ]
        msp.add_lwpolyline(pts)

    @staticmethod
    def _add_text_center_dxf(
        msp: Any,
        text: str,
        x: float,
        y: float,
        text_h: float,
    ) -> None:
        """Add centered text in DXF using MIDDLE_CENTER alignment."""
        if not text:
            return
        from ezdxf.enums import TextEntityAlignment

        txt = msp.add_text(str(text), dxfattribs={"height": text_h})
        txt.set_placement((x, y), align=TextEntityAlignment.MIDDLE_CENTER)

    def _insert_shape_block_dxf(
        self,
        msp: Any,
        zone: StirrupZone,
        x: float,
        y_top: float,
        width: float,
        height: float,
        cover_cm: float,
        hook_factor: float,
    ) -> None:
        """Insert TO block with L1/L2/L3 attributes in DXF modelspace."""
        doc = msp.doc
        if "TO" not in doc.blocks:
            return

        values = self._stirrup_shape_values(zone, cover_cm, hook_factor)
        insert_x = x + width * 0.5
        insert_y = y_top - height * 0.60
        bref = msp.add_blockref(
            "TO",
            (insert_x, insert_y),
            dxfattribs={
                "xscale": self.block_scale,
                "yscale": self.block_scale,
                "zscale": self.block_scale,
                "rotation": 0.0,
            },
        )
        try:
            bref.add_auto_attribs(values)
        except Exception:
            pass

    @staticmethod
    def _used_diameters(zones: list[StirrupZone]) -> list[int]:
        """Return the sorted list of unique rebar diameters actually used."""
        return sorted({int(z.diameter) for z in zones if z.diameter and z.total_length > 0})

    @staticmethod
    def _stirrup_table_headers(dias: list[int]) -> tuple[list[str], int]:
        """Build table headers: fixed columns + one column per used diameter."""
        fixed = [
            "Row", "POS", "Description", "Shape", "Zone", "Spc",
            "ZoneLen", "Cnt", "B", "H", "SingleL", "TotalL", "UnitW",
        ]
        headers = fixed + [f"T{d}" for d in dias]
        return headers, len(fixed)

    @staticmethod
    def _stirrup_row_values(
        i: int,
        zone: StirrupZone,
        dias: list[int],
        zone_type_map: dict[str, str],
    ) -> list[str]:
        """Build one data row's cell values, placing the weight in its size column."""
        values = [
            str(i),
            f"S{zone.pos:02d}",
            zone.description,
            "",
            zone_type_map.get(zone.zone_type, zone.zone_type),
            str(int(zone.spacing)) if zone.spacing else "",
            str(int(zone.zone_length)) if zone.zone_length else "",
            str(zone.count),
            str(int(zone.beam.width)) if zone.beam else "",
            str(int(zone.beam.height)) if zone.beam else "",
            f"{zone.single_length:.2f}",
            f"{zone.total_length:.2f}",
            f"{zone.unit_weight:.3f}",
        ]
        zone_dia = int(zone.diameter) if zone.diameter else None
        for dia in dias:
            values.append(f"{zone.total_weight:.2f}" if zone_dia == dia else "")
        return values

    @staticmethod
    def _diameter_summary_maps(
        summary: list[dict[str, Any]],
    ) -> tuple[dict[int, float], dict[int, float], float]:
        """Split ``summary_by_size()`` rows into per-diameter maps + grand total."""
        length_map: dict[int, float] = {}
        weight_map: dict[int, float] = {}
        grand_weight = 0.0
        for row in summary:
            size = row.get('Size (mm)')
            if size == 'TOTAL':
                grand_weight = float(row.get('Total Weight (kg)') or 0.0)
                continue
            if not isinstance(size, (int, float)):
                continue
            dia = int(size)
            length_map[dia] = float(row.get('Total Length (m)') or 0.0)
            weight_map[dia] = float(row.get('Total Weight (kg)') or 0.0)
        return length_map, weight_map, grand_weight

    def _draw_cell_acad(self, x: float, y_top: float, width: float, height: float) -> None:
        """Draw one rectangular table cell directly in AutoCAD."""
        pts = [
            x, y_top,
            x + width, y_top,
            x + width, y_top - height,
            x, y_top - height,
            x, y_top,
        ]
        self.msp.AddLightWeightPolyline(
            win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, pts)
        )

    def _add_text_center_acad(
        self,
        text: str,
        x: float,
        y: float,
        text_h: float,
        color: int | None = None,
    ) -> None:
        """Add centered text directly in AutoCAD."""
        if not text:
            return
        text_obj = self.msp.AddText(
            str(text),
            win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, 0)),
            text_h,
        )
        try:
            text_obj.Alignment = 1  # acAlignmentCenter
            text_obj.TextAlignmentPoint = win32com.client.VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, 0)
            )
        except Exception:
            pass
        if color:
            try:
                text_obj.Color = color
            except Exception:
                pass
        try:
            text_obj.Layer = self.doc.ActiveLayer.Name
        except Exception:
            pass

    def _draw_stirrup_summary_acad(
        self,
        extractor: StirrupFromDwg,
        dias: list[int],
        x0: float,
        y: float,
        col_widths: list[float],
        cell_h: float,
        text_h: float,
        fixed_col_count: int,
    ) -> None:
        """Draw the per-size TOTAL LENGTH / TOTAL WEIGHT / PERCENTAGE / GRAND TOTAL block."""
        if not dias:
            return
        length_map, weight_map, grand_weight = self._diameter_summary_maps(
            extractor.summary_by_size()
        )
        label_width = sum(col_widths[:fixed_col_count])
        dia_widths = col_widths[fixed_col_count:]

        def draw_label_row(label: str, dia_values: list[str]) -> None:
            nonlocal y
            self._draw_cell_acad(x0, y, label_width, cell_h)
            self._add_text_center_acad(
                label, x0 + label_width * 0.5, y - cell_h * 0.5, text_h,
                color=SUMMARY_FILL_COLOR,
            )
            x = x0 + label_width
            for width, value in zip(dia_widths, dia_values):
                self._draw_cell_acad(x, y, width, cell_h)
                self._add_text_center_acad(
                    value, x + width * 0.5, y - cell_h * 0.5, text_h,
                    color=SUMMARY_FILL_COLOR,
                )
                x += width
            y -= cell_h

        draw_label_row(
            "TOTAL LENGTH (m)", [f"{length_map.get(d, 0.0):.2f}" for d in dias]
        )
        draw_label_row(
            "TOTAL WEIGHT (Kg)", [f"{weight_map.get(d, 0.0):.2f}" for d in dias]
        )
        draw_label_row(
            "PERCENTAGE (%)",
            [
                f"{(weight_map.get(d, 0.0) / grand_weight * 100.0):.0f}%"
                if grand_weight else "0%"
                for d in dias
            ],
        )

        # Grand total: merged label cell + merged value cell spanning all sizes
        self._draw_cell_acad(x0, y, label_width, cell_h)
        self._add_text_center_acad(
            "GRAND TOTAL (Kg)", x0 + label_width * 0.5, y - cell_h * 0.5, text_h,
            color=SUMMARY_FILL_COLOR,
        )
        dia_total_width = sum(dia_widths)
        self._draw_cell_acad(x0 + label_width, y, dia_total_width, cell_h)
        self._add_text_center_acad(
            f"{grand_weight:.2f} Kg.",
            x0 + label_width + dia_total_width * 0.5,
            y - cell_h * 0.5,
            text_h,
        )

    def _draw_stirrup_summary_dxf(
        self,
        msp: Any,
        extractor: StirrupFromDwg,
        dias: list[int],
        x0: float,
        y: float,
        col_widths: list[float],
        cell_h: float,
        text_h: float,
        fixed_col_count: int,
    ) -> None:
        """DXF equivalent of :meth:`_draw_stirrup_summary_acad`."""
        if not dias:
            return
        length_map, weight_map, grand_weight = self._diameter_summary_maps(
            extractor.summary_by_size()
        )
        label_width = sum(col_widths[:fixed_col_count])
        dia_widths = col_widths[fixed_col_count:]

        def draw_label_row(label: str, dia_values: list[str]) -> None:
            nonlocal y
            self._draw_cell_dxf(msp, x0, y, label_width, cell_h)
            self._add_text_center_dxf(
                msp, label, x0 + label_width * 0.5, y - cell_h * 0.5, text_h
            )
            x = x0 + label_width
            for width, value in zip(dia_widths, dia_values):
                self._draw_cell_dxf(msp, x, y, width, cell_h)
                self._add_text_center_dxf(msp, value, x + width * 0.5, y - cell_h * 0.5, text_h)
                x += width
            y -= cell_h

        draw_label_row(
            "TOTAL LENGTH (m)", [f"{length_map.get(d, 0.0):.2f}" for d in dias]
        )
        draw_label_row(
            "TOTAL WEIGHT (Kg)", [f"{weight_map.get(d, 0.0):.2f}" for d in dias]
        )
        draw_label_row(
            "PERCENTAGE (%)",
            [
                f"{(weight_map.get(d, 0.0) / grand_weight * 100.0):.0f}%"
                if grand_weight else "0%"
                for d in dias
            ],
        )

        self._draw_cell_dxf(msp, x0, y, label_width, cell_h)
        self._add_text_center_dxf(
            msp, "GRAND TOTAL (Kg)", x0 + label_width * 0.5, y - cell_h * 0.5, text_h
        )
        dia_total_width = sum(dia_widths)
        self._draw_cell_dxf(msp, x0 + label_width, y, dia_total_width, cell_h)
        self._add_text_center_dxf(
            msp,
            f"{grand_weight:.2f} Kg.",
            x0 + label_width + dia_total_width * 0.5,
            y - cell_h * 0.5,
            text_h,
        )

    def draw_table_to_dxf(
        self,
        extractor: StirrupFromDwg,
        insert_point: tuple[float, float] = (0.0, 0.0),
        scale: float = 1.0,
        col_widths: list[float] | None = None,
        dia_col_width: float = DEFAULT_STIRRUP_TABLE_DIA_COL_WIDTH,
        cell_height: float = DEFAULT_STIRRUP_TABLE_CELL_H,
        text_height: float = DEFAULT_STIRRUP_TABLE_TEXT_H,
        min_text_height: float = DEFAULT_STIRRUP_TABLE_MIN_TEXT_H,
        output_path: str | Path | None = None,
        filename: str | None = None,
        open_file: bool = True,
    ) -> Path:
        """Write stirrup listofer into a DXF file (template copy + write + open)."""
        try:
            import ezdxf
        except ImportError as exc:
            raise ImportError(
                "ezdxf is required for DXF output. Install: pip install ezdxf"
            ) from exc

        readfile_fn = getattr(ezdxf, "readfile", None)
        if readfile_fn is None:
            from ezdxf import filemanagement as _filemanagement

            readfile_fn = _filemanagement.readfile

        new_fn = getattr(ezdxf, "new", None)
        if new_fn is None:
            from ezdxf import filemanagement as _filemanagement

            new_fn = _filemanagement.new

        out_path = self._resolve_output_dxf_path(output_path, filename)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        if self.template_path and self.template_path.exists():
            shutil.copy2(self.template_path, out_path)
            dxf_doc = readfile_fn(str(out_path))
        else:
            dxf_doc = new_fn("R2010")

        msp = dxf_doc.modelspace()
        x0, y0 = insert_point
        zones = extractor.stirrup_zones
        cover_cm = extractor.cover
        hook_factor = extractor.hook_factor

        dias = self._used_diameters(zones)
        headers, fixed_col_count = self._stirrup_table_headers(dias)

        cell_h = cell_height * scale
        width_base = col_widths or DEFAULT_STIRRUP_TABLE_COL_WIDTHS
        scaled_col_widths = [w * scale for w in width_base] + [
            dia_col_width * scale for _ in dias
        ]
        text_h = resolve_text_height(scale, text_height, min_text_height=min_text_height)

        y = y0
        x = x0
        for width, title in zip(scaled_col_widths, headers):
            self._draw_cell_dxf(msp, x, y, width, cell_h)
            self._add_text_center_dxf(msp, title, x + width * 0.5, y - cell_h * 0.5, text_h)
            x += width
        y -= cell_h

        zone_type_map = {"start": "S", "mid": "M", "end": "E"}
        for i, zone in enumerate(zones, 1):
            values = self._stirrup_row_values(i, zone, dias, zone_type_map)

            x = x0
            for col_idx, (width, value) in enumerate(zip(scaled_col_widths, values)):
                self._draw_cell_dxf(msp, x, y, width, cell_h)
                if col_idx == 3:
                    self._insert_shape_block_dxf(
                        msp,
                        zone,
                        x,
                        y,
                        width,
                        cell_h,
                        cover_cm,
                        hook_factor,
                    )
                else:
                    self._add_text_center_dxf(
                        msp,
                        value,
                        x + width * 0.5,
                        y - cell_h * 0.5,
                        text_h,
                    )
                x += width
            y -= cell_h

        self._draw_stirrup_summary_dxf(
            msp, extractor, dias, x0, y, scaled_col_widths, cell_h, text_h, fixed_col_count,
        )

        dxf_doc.saveas(str(out_path))
        if open_file and os.name == "nt":
            try:
                os.startfile(str(out_path))
            except OSError:
                pass

        print(f"DXF listofer saved: {out_path}")
        return out_path

    def draw_table(
        self,
        extractor: StirrupFromDwg,
        insert_point: tuple[float, float] = (0.0, 0.0),
        scale: float = 1.0,
        col_widths: list[float] | None = None,
        dia_col_width: float = DEFAULT_STIRRUP_TABLE_DIA_COL_WIDTH,
        cell_height: float = DEFAULT_STIRRUP_TABLE_CELL_H,
        text_height: float | None = None,
        text_height_factor: float = DEFAULT_STIRRUP_TABLE_TEXT_FACTOR,
    ) -> None:
        """Draw the stirrup schedule table at *insert_point*."""
        x0, y0 = insert_point
        zones = extractor.stirrup_zones
        cover_cm = extractor.cover
        hook_factor = extractor.hook_factor

        dias = self._used_diameters(zones)
        headers, fixed_col_count = self._stirrup_table_headers(dias)

        cell_h = cell_height * scale
        width_base = col_widths or DEFAULT_STIRRUP_TABLE_COL_WIDTHS
        scaled_col_widths = [w * scale for w in width_base] + [
            dia_col_width * scale for _ in dias
        ]
        row_text_h = resolve_text_height(
            scale, text_height, text_height_factor, cell_height,
            DEFAULT_STIRRUP_TABLE_MIN_TEXT_H,
        )

        # Header row
        y = y0
        self._draw_row(
            x0,
            y,
            scaled_col_widths,
            cell_h,
            headers,
            fill_color=HEADER_FILL_COLOR,
            text_height=row_text_h,
        )
        y -= cell_h

        # Data rows
        zone_type_map = {'start': 'S', 'mid': 'M', 'end': 'E'}
        for i, zone in enumerate(zones, 1):
            values = self._stirrup_row_values(i, zone, dias, zone_type_map)
            self._draw_row(
                x0,
                y,
                scaled_col_widths,
                cell_h,
                values,
                text_height=row_text_h,
                shape_col_idx=3,
                shape_callback=lambda cx, cy, cw, ch, z=zone: self._insert_shape_block(
                    z,
                    cx,
                    cy,
                    cw,
                    ch,
                    cover_cm,
                    hook_factor,
                ),
            )
            y -= cell_h

        # Summary block: TOTAL LENGTH / TOTAL WEIGHT / PERCENTAGE / GRAND TOTAL
        self._draw_stirrup_summary_acad(
            extractor, dias, x0, y, scaled_col_widths, cell_h, row_text_h, fixed_col_count,
        )

    def _draw_row(
        self,
        x0: float,
        y: float,
        col_widths: list[float],
        height: float,
        values: list[str],
        fill_color: int = 0,
        shape_col_idx: int | None = None,
        shape_callback: Any = None,
        text_height: float | None = None,
    ) -> None:
        """Draw one table row with text centered in each cell."""
        x = x0
        for idx, (width, value) in enumerate(zip(col_widths, values)):
            # Cell border — flat list of (x, y) pairs for LWPOLYLINE
            pts = [
                x, y,
                x + width, y,
                x + width, y - height,
                x, y - height,
                x, y,
            ]
            self.msp.AddLightWeightPolyline(
                win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, pts)
            )

            if shape_col_idx is not None and idx == shape_col_idx:
                if shape_callback is not None:
                    shape_callback(x, y, width, height)
                x += width
                continue

            # Text — use Left alignment with InsertionPoint for reliability
            if value:
                text_x = x + width / 2
                text_y = y - height / 2
                text_height_row = (
                    text_height
                    if text_height is not None
                    else max(
                        height * DEFAULT_STIRRUP_TABLE_TEXT_FACTOR,
                        DEFAULT_STIRRUP_TABLE_MIN_TEXT_H,
                    )
                )

                # Create text with Left alignment first (most reliable)
                text_obj = self.msp.AddText(
                    str(value),
                    win32com.client.VARIANT(
                        pythoncom.VT_ARRAY | pythoncom.VT_R8,
                        (text_x, text_y, 0)
                    ),
                    text_height_row,
                )

                # Set alignment to MiddleCenter (4 = acAlignmentCenter in some versions)
                # Note: In AutoCAD COM, Alignment values:
                # 0 = acAlignmentLeft, 1 = acAlignmentCenter, 2 = acAlignmentRight
                # 3 = acAlignmentAligned, 4 = acAlignmentMiddle, 5 = acAlignmentFit
                # 10 = acAlignmentMiddleCenter (most common for centered)
                try:
                    text_obj.Alignment = 1  # acAlignmentCenter
                    text_obj.TextAlignmentPoint = win32com.client.VARIANT(
                        pythoncom.VT_ARRAY | pythoncom.VT_R8,
                        (text_x, text_y, 0)
                    )
                except Exception:
                    # Fallback: just use InsertionPoint
                    pass

                # Set color (0 = ByLayer, use ByLayer for visibility)
                if fill_color:
                    try:
                        text_obj.Color = fill_color
                    except Exception:
                        pass

                # Ensure text is on current layer
                try:
                    text_obj.Layer = self.doc.ActiveLayer.Name
                except Exception:
                    pass

            x += width


# ---------------------------------------------------------------------------
#  Stirrup Shape Drawer
# ---------------------------------------------------------------------------

class StirrupShapeDrawer:
    """Draw stirrup U-shapes in AutoCAD for visual reference."""

    def __init__(self, doc: Any = None) -> None:
        if doc is None:
            self.acad = win32com.client.Dispatch("AutoCAD.Application")
            self.doc = self.acad.ActiveDocument
        else:
            self.doc = doc
        self.msp = self.doc.ModelSpace

    def draw_shape(
        self,
        zone: StirrupZone,
        beam: BeamDimensions | None = None,
        insert_point: tuple[float, float] = (0.0, 0.0),
        scale: float = 1.0,
        cover_cm: float = CONCRETE_COVER_CM,
        hook_factor: float = DEFAULT_HOOK_FACTOR,
    ) -> None:
        """Draw a closed stirrup U-shape at *insert_point*."""
        # Use zone's own beam dimensions if none provided
        if beam is None:
            beam = zone.beam

        x0, y0 = insert_point
        d_cm = zone.diameter / 10.0 * scale

        h_eff = (beam.height - 2 * cover_cm) * scale
        b_eff = (beam.width - 2 * cover_cm) * scale
        hook = (hook_factor * d_cm) * scale

        # U-shape with hooks
        pts = [
            (x0, y0),
            (x0, y0 + h_eff),
            (x0 + b_eff, y0 + h_eff),
            (x0 + b_eff, y0),
            (x0 + b_eff - hook, y0),
            (x0 + b_eff, y0 + hook),
            (x0 + b_eff, y0),
            (x0, y0),
            (x0 + hook, y0),
            (x0, y0 + hook),
            (x0, y0),
        ]

        flat = [c for p in pts for c in p]
        self.msp.AddLightWeightPolyline(
            win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, flat)
        )

        # Label
        label_y = y0 + h_eff + 5 * scale
        text_obj = self.msp.AddText(
            f"S{zone.pos:02d} - {zone.description}",
            win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8,
                                    (x0 + b_eff / 2, label_y, 0)),
            3 * scale,
        )
        text_obj.Alignment = 4
        text_obj.TextAlignmentPoint = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8, (x0 + b_eff / 2, label_y, 0))


# ---------------------------------------------------------------------------
#  Convenience / Demo
# ---------------------------------------------------------------------------

def demo():
    """Run a demonstration with sample data (no AutoCAD required)."""
    beam = BeamDimensions(width=60.0, height=70.0)
    extractor = StirrupFromDwg(beam_dims=beam, cover_cm=4.0, hook_factor=10.0)

    # Manually create sample zones (as if parsed from AutoCAD)
    zone1 = StirrupZone(
        pos=1, zone_type='start', diameter=12, spacing=15, zone_length=140,
        count=11, single_length=2.48, total_length=27.28,
        unit_weight=0.888, total_weight=24.22,
        has_add=True, add_count=1,
        beam=BeamDimensions(width=60, height=70),
    )
    zone2 = StirrupZone(
        pos=2, zone_type='mid', diameter=12, spacing=30, zone_length=458,
        count=16, single_length=2.48, total_length=39.68,
        unit_weight=0.888, total_weight=35.24,
        has_add=False,
        beam=BeamDimensions(width=60, height=70),
    )
    zone3 = StirrupZone(
        pos=3, zone_type='end', diameter=12, spacing=15, zone_length=140,
        count=11, single_length=2.48, total_length=27.28,
        unit_weight=0.888, total_weight=24.22,
        has_add=True, add_count=1,
        beam=BeamDimensions(width=60, height=70),
    )
    extractor.stirrup_zones = [zone1, zone2, zone3]

    print("=" * 60)
    print("STIRRUP SCHEDULE DEMO")
    print("=" * 60)
    print(f"Beam: {beam.width} x {beam.height} cm")
    print(f"Cover: {extractor.cover} cm")
    print(f"Hook factor: {extractor.hook_factor}d")
    print("-" * 60)

    for z in extractor.stirrup_zones:
        print(f"  POS S{z.pos:02d} | {z.zone_type:6s} | {z.description:12s} | "
              f"Len={z.zone_length:3.0f}cm | Spc={z.spacing:2.0f}cm | "
              f"Cnt={z.count:2d} | Single={z.single_length:.2f}m | "
              f"Wt={z.total_weight:6.2f} kg | Beam={z.beam.width}x{z.beam.height}")

    print("-" * 60)
    summary = extractor.summary_by_size()
    for row in summary:
        print(f"  {row}")
    print("=" * 60)

    # Verify single_length calculation
    print("\n--- Verification ---")
    test_len = calculate_stirrup_length(12, beam, 4.0, 10.0)
    print(f"calculate_stirrup_length(12mm, 60x70cm, cover=4cm, hook=10d) = {test_len} m")
    test_wt = calculate_unit_weight(12)
    print(f"calculate_unit_weight(12mm) = {test_wt} kg/m")

    # Excel export
    try:
        export_to_excel(extractor, "stirrup_schedule_demo.xlsx")
        print("\n✅ Excel saved: stirrup_schedule_demo.xlsx")
    except ImportError as e:
        print(f"\n⚠️  Excel export skipped: {e}")

    return extractor


if __name__ == "__main__":
    demo()
